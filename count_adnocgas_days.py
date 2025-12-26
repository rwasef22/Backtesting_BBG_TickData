import openpyxl
from datetime import datetime

print("Loading TickData.xlsx...")
wb = openpyxl.load_workbook('data/raw/TickData.xlsx', read_only=False, data_only=True)

print(f"\nAll sheets in workbook:")
for sheet_name in wb.sheetnames:
    print(f"  - {sheet_name}")

# Look for ADNOCGAS sheet
adnocgas_sheet = None
for sheet_name in wb.sheetnames:
    if 'ADNOCGAS' in sheet_name.upper():
        adnocgas_sheet = sheet_name
        break

if not adnocgas_sheet:
    print("\nNo ADNOCGAS sheet found!")
    wb.close()
    exit(1)

print(f"\nFound ADNOCGAS sheet: {adnocgas_sheet}")
ws = wb[adnocgas_sheet]

print(f"Max row: {ws.max_row:,}")
print(f"Max col: {ws.max_column}")

# Read first few rows to understand structure
print("\nFirst 5 rows:")
for i, row in enumerate(ws.iter_rows(min_row=1, max_row=5, values_only=True), start=1):
    print(f"Row {i}: {row}")

# Find the header row (should be row 3 based on pattern)
header_row = 3
header = [cell.value for cell in ws[header_row]]
print(f"\nHeader (row {header_row}): {header}")

# Count unique dates
dates_col = header.index('Dates')
adnocgas_dates = set()

print("\nProcessing rows...")
for i, row in enumerate(ws.iter_rows(min_row=header_row+1, values_only=True), start=header_row+1):
    if i % 10000 == 0:
        print(f"  Processed {i-header_row:,} rows, found {len(adnocgas_dates)} unique dates...")
    
    date_val = row[dates_col]
    if isinstance(date_val, datetime):
        adnocgas_dates.add(date_val.date())

wb.close()

print(f"\n{'='*60}")
print(f"ADNOCGAS Market Trading Days Analysis")
print(f"{'='*60}")
print(f"Total data rows: {ws.max_row - header_row:,}")
print(f"Unique trading dates: {len(adnocgas_dates)}")

if adnocgas_dates:
    sorted_dates = sorted(adnocgas_dates)
    print(f"\nDate range:")
    print(f"  First date: {sorted_dates[0]}")
    print(f"  Last date: {sorted_dates[-1]}")
    print(f"\nAll {len(sorted_dates)} trading dates:")
    for date in sorted_dates:
        print(f"  {date}")
