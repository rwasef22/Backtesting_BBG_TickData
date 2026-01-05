"""Minimal Excel to Parquet converter - direct openpyxl usage."""
import sys
import os
import time
from pathlib import Path
from datetime import datetime

# Don't import from project - keep this standalone
import pandas as pd
from openpyxl import load_workbook


def convert_sheet_to_parquet(wb, sheet_name, output_dir, header_row=3):
    """Convert one sheet to Parquet."""
    ws = wb[sheet_name]
    
    # Read all rows
    rows = list(ws.iter_rows(values_only=True))
    
    # Get header
    if len(rows) < header_row:
        return None
        
    header = rows[header_row - 1]
    columns = [str(c).strip().lower() if c is not None else f'col_{i}' 
               for i, c in enumerate(header)]
    
    # Get data
    data_rows = rows[header_row:]
    
    # Create DataFrame
    df = pd.DataFrame(data_rows, columns=columns)
    
    # Handle 'dates' column (rename to timestamp)
    if 'dates' in df.columns:
        df['timestamp'] = pd.to_datetime(df['dates'], errors='coerce')
        df = df.drop(['dates'], axis=1)
    # Combine date and time if separate
    elif 'date' in df.columns and 'time' in df.columns:
        df['timestamp'] = pd.to_datetime(
            df['date'].astype(str) + ' ' + df['time'].astype(str),
            errors='coerce'
        )
        df = df.drop(['date', 'time'], axis=1)
    
    # Handle 'size' column (rename to volume)
    if 'size' in df.columns:
        df['volume'] = df['size']
        df = df.drop(['size'], axis=1)
    
    # Keep only required columns
    required = ['timestamp', 'type', 'price', 'volume']
    available = [c for c in required if c in df.columns]
    
    if len(available) < 4:
        print(f"  Warning: Missing columns. Have: {list(df.columns)}")
        return None
    
    df = df[available]
    df = df.dropna(subset=['timestamp', 'price'])
    
    # Get security name
    security = sheet_name.replace(' UH Equity', '').replace(' DH Equity', '').lower()
    
    # Write Parquet
    output_path = Path(output_dir)
    output_path.mkdir(parents=True, exist_ok=True)
    parquet_file = output_path / f"{security}.parquet"
    
    df.to_parquet(parquet_file, compression='snappy', index=False)
    
    return len(df), parquet_file


def main():
    excel_path = 'data/raw/TickData.xlsx'
    output_dir = 'data/parquet'
    
    print("="*60)
    print("MINIMAL EXCEL TO PARQUET CONVERTER")
    print("="*60)
    print(f"Loading: {excel_path}")
    
    start = time.time()
    
    # Load workbook (not read_only for reliability)
    wb = load_workbook(excel_path, read_only=False, data_only=True)
    print(f"Loaded in {time.time()-start:.1f}s")
    print(f"Sheets: {len(wb.sheetnames)}")
    print()
    
    converted = 0
    total_rows = 0
    
    for i, sheet_name in enumerate(wb.sheetnames, 1):
        print(f"[{i}/{len(wb.sheetnames)}] {sheet_name}... ", end='', flush=True)
        
        try:
            result = convert_sheet_to_parquet(wb, sheet_name, output_dir)
            if result:
                rows, pf = result
                size_mb = pf.stat().st_size / (1024*1024)
                print(f"{rows:,} rows, {size_mb:.1f} MB")
                converted += 1
                total_rows += rows
            else:
                print("skipped")
        except Exception as e:
            print(f"ERROR: {e}")
    
    wb.close()
    
    elapsed = time.time() - start
    print()
    print("="*60)
    print(f"DONE: {converted} sheets, {total_rows:,} rows in {elapsed:.1f}s")
    print("="*60)


if __name__ == '__main__':
    main()
