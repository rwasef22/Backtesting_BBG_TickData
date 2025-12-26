"""
Compare market trading days across all securities.
"""
import openpyxl
from datetime import datetime

wb = openpyxl.load_workbook('data/raw/TickData.xlsx', read_only=True, data_only=True)

results = []

for sheet_name in wb.sheetnames:
    security = sheet_name.replace(' UH Equity', '').replace(' DH Equity', '')
    ws = wb[sheet_name]
    
    market_dates = set()
    
    # Count days with any market activity
    for row in ws.iter_rows(min_row=4, max_row=200000, values_only=True):
        if not row[0]:
            continue
        
        dt = row[0]
        if isinstance(dt, datetime):
            market_dates.add(dt.date().isoformat())
    
    results.append({
        'Security': security,
        'Market Days': len(market_dates)
    })

wb.close()

# Sort by market days
results.sort(key=lambda x: x['Market Days'])

print("\nMARKET TRADING DAYS BY SECURITY")
print("="*50)
for r in results:
    print(f"{r['Security']:15} {r['Market Days']:3} days")
print("="*50)
