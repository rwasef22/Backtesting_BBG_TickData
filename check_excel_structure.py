import openpyxl

print("Loading TickData.xlsx...")
wb = openpyxl.load_workbook('data/raw/TickData.xlsx', read_only=True, data_only=True)
ws = wb.active

print(f"\nSheet name: {ws.title}")
print(f"Max row: {ws.max_row}")
print(f"Max col: {ws.max_column}")

print("\nFirst 5 rows:")
for i, row in enumerate(ws.iter_rows(min_row=1, max_row=5, values_only=True), start=1):
    print(f"Row {i}: {row}")

wb.close()
