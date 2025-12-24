#!/usr/bin/env python
import openpyxl

wb = openpyxl.load_workbook('data/raw/TickData.xlsx', read_only=True, data_only=True)
print(f"Total sheets in workbook: {len(wb.sheetnames)}")
print(f"Sheet names:")
for i, name in enumerate(wb.sheetnames, 1):
    ws = wb[name]
    row_count = ws.max_row - 1 if ws.max_row else 0
    print(f"  {i}. {name} ({row_count} rows)")
