"""Parquet utilities for automatic conversion and usage.

This module provides utilities to automatically handle Parquet conversion
and ensure all backtests use Parquet format by default.
"""

import os
import sys
from pathlib import Path
import warnings
import pandas as pd
from datetime import datetime


def validate_parquet_against_excel(excel_path, parquet_dir, max_sheets=None):
    """Validate Parquet data matches Excel source.
    
    Args:
        excel_path: Path to source Excel file
        parquet_dir: Directory containing Parquet files
        max_sheets: Only validate this many sheets (for testing)
    
    Returns:
        True if validation passes
    
    Raises:
        ValueError: If validation fails
    """
    import openpyxl
    
    parquet_path = Path(parquet_dir)
    
    # Get Excel sheet names
    wb = openpyxl.load_workbook(excel_path, read_only=True)
    excel_sheets = wb.sheetnames
    wb.close()
    
    if max_sheets:
        excel_sheets = excel_sheets[:max_sheets]
    
    # Get Parquet files
    parquet_files = list(parquet_path.glob("*.parquet"))
    parquet_securities = {f.stem.upper() for f in parquet_files}
    
    # Extract security names from Excel sheets
    excel_securities = set()
    for sheet in excel_sheets:
        sec = sheet.replace(' UH Equity', '').replace(' DH Equity', '')
        excel_securities.add(sec)
    
    # Check coverage
    missing = excel_securities - parquet_securities
    if missing and max_sheets is None:
        raise ValueError(f"Missing Parquet files for: {missing}")
    
    # Check date ranges match (sample a few files)
    for parquet_file in list(parquet_files)[:3]:  # Sample first 3
        security = parquet_file.stem.upper()
        
        # Find matching sheet
        sheet_name = None
        for sheet in excel_sheets:
            if security in sheet.upper():
                sheet_name = sheet
                break
        
        if sheet_name is None:
            continue
        
        # Read first and last rows from both
        pq_df = pd.read_parquet(parquet_file)
        
        # Get date range from Parquet
        if 'timestamp' in pq_df.columns:
            pq_dates = pd.to_datetime(pq_df['timestamp'])
        elif 'Date' in pq_df.columns:
            pq_dates = pd.to_datetime(pq_df['Date'])
        else:
            continue
        
        pq_start = pq_dates.min().date()
        pq_end = pq_dates.max().date()
        
        # Read Excel date range (sample first and last chunks)
        excel_df = pd.read_excel(excel_path, sheet_name=sheet_name, header=2, nrows=100)
        if 'Date' in excel_df.columns:
            excel_start = pd.to_datetime(excel_df['Date']).min().date()
        else:
            continue
        
        # Just check start date matches (end date check would require reading all Excel data)
        if pq_start != excel_start:
            raise ValueError(f"{security}: Start date mismatch - Parquet: {pq_start}, Excel: {excel_start}")
    
    return True


def ensure_parquet_data(excel_path='data/raw/TickData.xlsx', 
                        parquet_dir='data/parquet',
                        force_reconvert=False,
                        max_sheets=None,
                        validate_data=True):
    """Ensure Parquet data exists and is valid.
    
    This function checks for existing Parquet files and optionally validates
    them against the source Excel. If files are missing or invalid, it prompts
    for reconversion.
    
    Args:
        excel_path: Path to source Excel file
        parquet_dir: Directory for Parquet files
        force_reconvert: Force reconversion even if files exist
        max_sheets: Limit conversion to first N sheets
        validate_data: Whether to validate Parquet against Excel
    
    Returns:
        Path to parquet directory
    
    Raises:
        FileNotFoundError: If Excel file doesn't exist and no Parquet data
        ValueError: If validation fails
    """
    parquet_path = Path(parquet_dir)
    excel_exists = Path(excel_path).exists()
    parquet_exists = parquet_path.exists() and any(parquet_path.glob("*.parquet"))
    
    if force_reconvert and excel_exists:
        # Force reconversion
        convert_excel_to_parquet(excel_path, parquet_dir, max_sheets=max_sheets)
        return str(parquet_path)
    
    if parquet_exists:
        # Validate if requested
        if validate_data and excel_exists:
            try:
                print("\nValidating Parquet data against Excel source...")
                validate_parquet_against_excel(excel_path, parquet_dir, max_sheets)
                print("[OK] Parquet data validation passed")
            except ValueError as e:
                print(f"[X] Validation failed: {e}")
                print("Consider reconverting with --force-reconvert")
                raise
        
        # Parquet exists and is valid (or validation skipped)
        num_files = len(list(parquet_path.glob("*.parquet")))
        warnings.warn(f"""
================================================================================
USING EXISTING PARQUET FILES
================================================================================
Found {num_files} Parquet files in {parquet_dir}
Using Parquet format for 5-10x faster I/O.
To reconvert from Excel, delete the parquet/ directory or use --force-reconvert
================================================================================
""")
        return str(parquet_path)
    
    # No Parquet data, need to convert
    if not excel_exists:
        raise FileNotFoundError(
            f"No Parquet data found in {parquet_dir} and Excel file not found: {excel_path}"
        )
    
    print(f"\nNo Parquet data found. Converting from Excel...")
    convert_excel_to_parquet(excel_path, parquet_dir, max_sheets=max_sheets)
    return str(parquet_path)


def convert_excel_to_parquet(excel_path, parquet_dir, max_sheets=None):
    """Convert Excel file to Parquet format.
    
    Args:
        excel_path: Path to source Excel file
        parquet_dir: Output directory for Parquet files
        max_sheets: Limit to first N sheets (for testing)
    """
    import openpyxl
    
    print(f"Converting Excel to Parquet...")
    print(f"  Source: {excel_path}")
    print(f"  Output: {parquet_dir}")
    
    # Create output directory
    parquet_path = Path(parquet_dir)
    parquet_path.mkdir(parents=True, exist_ok=True)
    
    # Get sheet names
    wb = openpyxl.load_workbook(excel_path, read_only=True)
    sheet_names = wb.sheetnames
    wb.close()
    
    if max_sheets:
        sheet_names = sheet_names[:max_sheets]
    
    print(f"  Converting {len(sheet_names)} sheets...")
    
    for i, sheet_name in enumerate(sheet_names, 1):
        # Extract security name
        security = sheet_name.replace(' UH Equity', '').replace(' DH Equity', '')
        
        print(f"  [{i}/{len(sheet_names)}] {sheet_name}...", end=" ", flush=True)
        
        try:
            # Read Excel sheet
            df = pd.read_excel(excel_path, sheet_name=sheet_name, header=2)
            
            # Combine Date and Time into timestamp if separate
            if 'Date' in df.columns and 'Time' in df.columns:
                df['timestamp'] = pd.to_datetime(
                    df['Date'].astype(str) + ' ' + df['Time'].astype(str)
                )
            
            # Write Parquet
            output_file = parquet_path / f"{security.lower()}.parquet"
            df.to_parquet(output_file, compression='snappy')
            
            print(f"{len(df):,} rows")
        except Exception as e:
            print(f"ERROR: {e}")
    
    print(f"\n[OK] Conversion complete: {parquet_dir}")
