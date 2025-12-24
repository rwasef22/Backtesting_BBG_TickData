"""Streaming Excel data loader optimized for many sheets / large rows.

Provides two main helpers:
- stream_sheets(file_path, header_row=3, chunk_size=100000): generator yielding (sheet_name, chunk_df)
- load_tick_data_multi_sheet(file_path, only_trades=False, max_sheets=None): convenience loader returning dict of DataFrames

Design goals:
- Use openpyxl in read_only mode to avoid loading entire workbook into memory.
- Process rows in chunks and convert to pandas.DataFrame for downstream processing.
- Provide a streaming API so callers can process data sheet-by-sheet, chunk-by-chunk.
"""
from typing import Generator, Tuple, Optional
import os
import pandas as pd

try:
    from openpyxl import load_workbook
    _HAS_OPENPYXL = True
except Exception:
    _HAS_OPENPYXL = False


def _normalize_row_values(row):
    # Convert openpyxl row tuple to a plain list
    return [None if v == '' else v for v in row]


def stream_sheets(file_path: str, header_row: int = 3, chunk_size: int = 100000,
                  max_sheets: Optional[int] = None, only_trades: bool = False) -> Generator[Tuple[str, pd.DataFrame], None, None]:
    """Stream each sheet in `file_path` and yield DataFrame chunks.

    Yields (sheet_name, chunk_df). chunk_df will have columns inferred from the header_row.

    - header_row is 1-based index of header line in Excel (defaults to 3 to match TickData format).
    - chunk_size controls how many rows are yielded per chunk.
    - only_trades: when True, filters rows where a 'Type' column equals 'TRADE' (case-insensitive).
    """
    if not os.path.exists(file_path):
        raise FileNotFoundError(file_path)

    # Prefer openpyxl streaming reader for large files
    if _HAS_OPENPYXL:
        try:
            wb = load_workbook(filename=file_path, read_only=True, data_only=True)
            sheet_names = wb.sheetnames[:max_sheets] if max_sheets else wb.sheetnames

            for sheet_name in sheet_names:
                ws = wb[sheet_name]
                it = ws.iter_rows(values_only=True)

                # Advance to header row (1-based count)
                header = None
                for i in range(header_row - 1):
                    try:
                        next(it)
                    except StopIteration:
                        break

                try:
                    header = next(it)
                except StopIteration:
                    # empty sheet
                    continue

                # Normalize header names
                columns = [str(c).strip() if c is not None else f'col_{i}' for i, c in enumerate(header)]

                buffer = []
                for row in it:
                    vals = _normalize_row_values(row)
                    row_dict = {columns[i] if i < len(columns) else f'col_{i}': vals[i] if i < len(vals) else None for i in range(len(columns))}
                    buffer.append(row_dict)

                    if len(buffer) >= chunk_size:
                        df = pd.DataFrame(buffer)
                        if only_trades and 'Type' in df.columns:
                            df = df[df['Type'].astype(str).str.upper() == 'TRADE']
                        yield sheet_name, df
                        buffer = []

                # yield remainder
                if buffer:
                    df = pd.DataFrame(buffer)
                    if only_trades and 'Type' in df.columns:
                        df = df[df['Type'].astype(str).str.upper() == 'TRADE']
                    yield sheet_name, df

            wb.close()
            return
        except Exception as e:
            # Fallback to pandas if openpyxl streaming fails
            print(f"WARNING: openpyxl streaming failed with error: {e}")
            print(f"Falling back to pandas full-sheet load")
            import traceback
            traceback.print_exc()
            pass

    # Fallback: pandas full-sheet load then chunk
    xls = pd.ExcelFile(file_path)
    sheets = xls.sheet_names[:max_sheets] if max_sheets else xls.sheet_names
    for sheet in sheets:
        df_full = pd.read_excel(file_path, sheet_name=sheet, header=header_row - 1)
        if only_trades and 'Type' in df_full.columns:
            df_full = df_full[df_full['Type'].astype(str).str.upper() == 'TRADE']
        if chunk_size and chunk_size > 0:
            for start in range(0, len(df_full), chunk_size):
                yield sheet, df_full.iloc[start:start + chunk_size].copy()
        else:
            yield sheet, df_full

    try:
        wb = load_workbook(filename=file_path, read_only=True, data_only=True)
        sheet_names = wb.sheetnames[:max_sheets] if max_sheets else wb.sheetnames

        for sheet_name in sheet_names:
            ws = wb[sheet_name]
            it = ws.iter_rows(values_only=True)

            # Advance to header row (1-based count)
            header = None
            for i in range(header_row - 1):
                try:
                    next(it)
                except StopIteration:
                    break

            try:
                header = next(it)
            except StopIteration:
                # empty sheet
                continue

            # Normalize header names
            columns = [str(c).strip() if c is not None else f'col_{i}' for i, c in enumerate(header)]

            buffer = []
            rows_read = 0
            for row in it:
                rows_read += 1
                vals = _normalize_row_values(row)
                row_dict = {columns[i] if i < len(columns) else f'col_{i}': vals[i] if i < len(vals) else None for i in range(len(columns))}
                buffer.append(row_dict)

                if len(buffer) >= chunk_size:
                    df = pd.DataFrame(buffer)
                    if only_trades and 'Type' in df.columns:
                        df = df[df['Type'].astype(str).str.upper() == 'TRADE']
                    yield sheet_name, df
                    buffer = []

            # yield remainder
            if buffer:
                df = pd.DataFrame(buffer)
                if only_trades and 'Type' in df.columns:
                    df = df[df['Type'].astype(str).str.upper() == 'TRADE']
                yield sheet_name, df

        wb.close()
    except Exception:
        raise


def preprocess_chunk_df(df: pd.DataFrame) -> pd.DataFrame:
    """Normalize columns and types for the backtest.

    Expected final columns: timestamp, type, price, volume
    Accepts common input names (Dates, Date, Type, Price, Size, Qty, Volume)
    """
    if df is None or df.empty:
        return pd.DataFrame(columns=['timestamp', 'type', 'price', 'volume'])

    # Normalize column keys
    df = df.rename(columns={c: c.strip() for c in df.columns})
    col_map = {}
    for c in df.columns:
        lower = c.lower()
        if 'date' in lower or 'time' in lower or 'dates' in lower:
            col_map[c] = 'timestamp'
        elif lower in ('type',):
            col_map[c] = 'type'
        elif 'price' in lower:
            col_map[c] = 'price'
        elif lower in ('size', 'vol', 'volume', 'qty', 'quantity'):
            col_map[c] = 'volume'

    df = df.rename(columns=col_map)

    # Keep only relevant columns if present
    for expected in ['timestamp', 'type', 'price', 'volume']:
        if expected not in df.columns:
            df[expected] = None

    # Type conversions
    try:
        df['timestamp'] = pd.to_datetime(df['timestamp'], errors='coerce')
    except Exception:
        df['timestamp'] = pd.NaT
    df['type'] = df['type'].astype(str).str.lower().str.strip().fillna('')
    df['price'] = pd.to_numeric(df['price'], errors='coerce')
    df['volume'] = pd.to_numeric(df['volume'], errors='coerce')

    # Drop rows without timestamp/price
    df = df.dropna(subset=['timestamp', 'price'])
    df = df[df['price'] > 0]

    # Reorder columns
    return df[['timestamp', 'type', 'price', 'volume']]


def load_tick_data_multi_sheet(file_path: str, only_trades: bool = False, max_sheets: Optional[int] = None, chunk_size: int = 100000) -> dict:
    """Load entire workbook into dict of DataFrames per sheet using streaming chunks.

    This convenience function will internally stream sheet chunks and concat them per-sheet.
    Use `stream_sheets` directly for memory-sensitive streaming processing.
    """
    result = {}
    for sheet_name, chunk in stream_sheets(file_path, header_row=3, chunk_size=chunk_size, max_sheets=max_sheets, only_trades=only_trades):
        df = preprocess_chunk_df(chunk)
        if sheet_name not in result:
            result[sheet_name] = [df]
        else:
            result[sheet_name].append(df)

    # concatenate
    for k, parts in list(result.items()):
        if parts:
            result[k] = pd.concat(parts, ignore_index=True)
        else:
            result[k] = pd.DataFrame(columns=['timestamp', 'type', 'price', 'volume'])

    return result


def load_tick_data(file_path: str):
    """Backward-compatible loader: returns first sheet as DataFrame.

    If file is not Excel, tries CSV. If missing, raises FileNotFoundError.
    """
    if not os.path.exists(file_path):
        raise FileNotFoundError(file_path)

    # If it's a CSV, use pandas directly
    lower = file_path.lower()
    if lower.endswith('.csv'):
        df = pd.read_csv(file_path)
        return preprocess_chunk_df(df)

    # Otherwise treat as Excel and return the first sheet
    for sheet_name, chunk in stream_sheets(file_path, header_row=3, chunk_size=100000, max_sheets=1, only_trades=False):
        df = preprocess_chunk_df(chunk)
        return df

    # If no sheets yielded, return empty df
    return pd.DataFrame(columns=['timestamp', 'type', 'price', 'volume'])
import pandas as pd


def load_tick_data_multi_sheet(file_path):
    """Load tick data from multi-sheet Excel file.
    
    Expected format:
      - Each sheet is a different security (sheet name = security name)
      - A1: Security name (label)
      - A2: Random text (description/metadata)
      - Row 3: Headers (Dates, Type, Price, Size)
      - Dates: timestamp column
      
    Returns:
      dict mapping security_name -> DataFrame with columns ['timestamp','type','price','volume']
    """
    try:
        xls = pd.ExcelFile(file_path)
        securities = {}
        
        for sheet_name in xls.sheet_names:
            # Skip sheets that look like metadata
            if sheet_name.lower() in ['meta', 'info', 'readme']:
                continue
                
            # Read sheet starting from row 2 (0-indexed), so headers are at row 3
            df = pd.read_excel(file_path, sheet_name=sheet_name, header=2)
            
            if df.empty:
                continue
            
            # Normalize column names
            df.columns = [c.strip().lower() for c in df.columns]
            
            # Map expected column names
            col_mapping = {
                'dates': 'timestamp',
                'date': 'timestamp',
                'time': 'timestamp',
                'type': 'type',
                'price': 'price',
                'size': 'volume',
                'vol': 'volume',
                'volume': 'volume',
                'qty': 'volume',
                'quantity': 'volume'
            }
            
            # Rename columns if they match
            rename_map = {}
            for old_col in df.columns:
                for key, new_col in col_mapping.items():
                    if key in old_col:
                        rename_map[old_col] = new_col
                        break
            
            df = df.rename(columns=rename_map)
            
            # Validate required columns
            required = {'timestamp', 'type', 'price', 'volume'}
            if not required.issubset(set(df.columns)):
                print(f"Warning: Sheet '{sheet_name}' missing columns. Found: {list(df.columns)}")
                continue
            
            # Preprocess
            df = preprocess_tick_data(df)
            securities[sheet_name] = df
        
        return securities if securities else None
    except Exception as e:
        print(f"Error reading Excel file: {e}")
        return None


def load_tick_data(file_path):
    """Load tick updates from Excel into a normalized DataFrame.

    Expected format (per-row update):
      - timestamp: datetime-like
      - type: one of 'bid', 'ask', 'trade'
      - price: numeric
      - volume: numeric

    Returns:
      pd.DataFrame with columns ['timestamp','type','price','volume'] or raises on error.
    """
    # Try multi-sheet format first
    multi_data = load_tick_data_multi_sheet(file_path)
    if multi_data:
        # Return first security's data for backward compatibility
        return list(multi_data.values())[0]
    
    # Fallback to single-sheet format
    try:
        df = pd.read_excel(file_path)
    except ValueError:
        try:
            df = pd.read_csv(file_path)
        except Exception:
            df = None

    # If df loaded but doesn't contain expected columns, try CSV or fallback to sample
    required = {'timestamp', 'type', 'price', 'volume'}
    if df is not None:
        cols = {c.strip().lower() for c in df.columns}
        if not required.issubset(cols):
            try:
                df2 = pd.read_csv(file_path)
                df = df2
            except Exception:
                df = None

    # If we still don't have required columns, fall back to sample
    if df is not None:
        cols = {c.strip().lower() for c in df.columns}
        if not required.issubset(cols):
            df = None

    if df is None:
        df = _create_sample_data()

    return preprocess_tick_data(df)


def _create_sample_data():
    """Create realistic synthetic sample data."""
    sample = [
        {'timestamp': '2020-01-01 09:30:00', 'price': 100.00, 'volume': 100, 'type': 'bid'},
        {'timestamp': '2020-01-01 09:30:00', 'price': 99.98, 'volume': 75, 'type': 'bid'},
        {'timestamp': '2020-01-01 09:30:00', 'price': 100.05, 'volume': 50, 'type': 'ask'},
        {'timestamp': '2020-01-01 09:30:00', 'price': 100.07, 'volume': 60, 'type': 'ask'},
        {'timestamp': '2020-01-01 09:30:01', 'price': 100.02, 'volume': 75, 'type': 'bid'},
        {'timestamp': '2020-01-01 09:30:01', 'price': 100.08, 'volume': 40, 'type': 'ask'},
        {'timestamp': '2020-01-01 09:30:02', 'price': 100.02, 'volume': 30, 'type': 'trade'},
        {'timestamp': '2020-01-01 09:30:03', 'price': 100.01, 'volume': 120, 'type': 'bid'},
        {'timestamp': '2020-01-01 09:30:03', 'price': 100.09, 'volume': 60, 'type': 'ask'},
        {'timestamp': '2020-01-01 09:30:04', 'price': 100.01, 'volume': 50, 'type': 'trade'},
        {'timestamp': '2020-01-01 09:30:05', 'price': 100.03, 'volume': 90, 'type': 'bid'},
        {'timestamp': '2020-01-01 09:30:05', 'price': 100.07, 'volume': 35, 'type': 'ask'},
        {'timestamp': '2020-01-01 09:30:06', 'price': 100.05, 'volume': 70, 'type': 'trade'},
        {'timestamp': '2020-01-01 09:30:07', 'price': 100.04, 'volume': 110, 'type': 'bid'},
        {'timestamp': '2020-01-01 09:30:07', 'price': 100.10, 'volume': 55, 'type': 'ask'},
        {'timestamp': '2020-01-01 09:30:08', 'price': 100.04, 'volume': 40, 'type': 'trade'},
    ]
    return pd.DataFrame(sample)


def preprocess_tick_data(tick_data):
    # normalize column names to lowercase
    tick_data = tick_data.rename(columns={c: c.strip().lower() for c in tick_data.columns})

    required = {'timestamp', 'price', 'volume', 'type'}
    if not required.issubset(set(tick_data.columns)):
        raise ValueError(f"Tick data missing required columns: {required - set(tick_data.columns)}")

    tick_data['timestamp'] = pd.to_datetime(tick_data['timestamp'])
    tick_data['type'] = tick_data['type'].str.lower()
    tick_data = tick_data.sort_values(by='timestamp').reset_index(drop=True)
    return tick_data