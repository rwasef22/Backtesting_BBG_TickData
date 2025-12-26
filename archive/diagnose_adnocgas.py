"""
Diagnose ADNOCGAS liquidity issues - check typical bid/ask sizes.
"""
import openpyxl
from datetime import datetime, time
from collections import defaultdict

wb = openpyxl.load_workbook('data/raw/TickData.xlsx', read_only=True, data_only=True)
ws = wb['ADNOCGAS UH Equity']

print("Analyzing ADNOCGAS liquidity...")

# Collect bid/ask sizes during trading hours (10:05-14:45)
bid_sizes = []
ask_sizes = []
daily_samples = defaultdict(lambda: {'bids': [], 'asks': []})

row_count = 0
for row in ws.iter_rows(min_row=4, max_row=50000, values_only=True):
    row_count += 1
    if not row[0] or not row[1]:
        continue
    
    dt = row[0]
    if not isinstance(dt, datetime):
        continue
    
    # Skip silent period and auctions
    t = dt.time()
    if t < time(10, 5) or t > time(14, 45):
        continue
    
    row_type = row[1]
    size = row[3] if len(row) > 3 and row[3] else 0
    
    if size <= 0:
        continue
    
    date_str = dt.date().isoformat()
    
    if row_type == 'BID':
        bid_sizes.append(size)
        daily_samples[date_str]['bids'].append(size)
    elif row_type == 'ASK':
        ask_sizes.append(size)
        daily_samples[date_str]['asks'].append(size)

print(f"\nProcessed {row_count} rows")
print(f"Total bid updates: {len(bid_sizes)}")
print(f"Total ask updates: {len(ask_sizes)}")

if bid_sizes:
    print(f"\nBid size statistics:")
    print(f"  Average: {sum(bid_sizes)/len(bid_sizes):,.0f}")
    print(f"  Median: {sorted(bid_sizes)[len(bid_sizes)//2]:,.0f}")
    print(f"  Min: {min(bid_sizes):,.0f}")
    print(f"  Max: {max(bid_sizes):,.0f}")
    print(f"  10th percentile: {sorted(bid_sizes)[len(bid_sizes)//10]:,.0f}")
    print(f"  90th percentile: {sorted(bid_sizes)[len(bid_sizes)*9//10]:,.0f}")

if ask_sizes:
    print(f"\nAsk size statistics:")
    print(f"  Average: {sum(ask_sizes)/len(ask_sizes):,.0f}")
    print(f"  Median: {sorted(ask_sizes)[len(ask_sizes)//2]:,.0f}")
    print(f"  Min: {min(ask_sizes):,.0f}")
    print(f"  Max: {max(ask_sizes):,.0f}")
    print(f"  10th percentile: {sorted(ask_sizes)[len(ask_sizes)//10]:,.0f}")
    print(f"  90th percentile: {sorted(ask_sizes)[len(ask_sizes)*9//10]:,.0f}")

# Check typical price to calculate liquidity in currency
print(f"\nChecking typical prices...")
prices = []
for row in ws.iter_rows(min_row=4, max_row=10000, values_only=True):
    if not row[0] or not row[2]:
        continue
    price = row[2]
    if price and price > 0:
        prices.append(price)
        if len(prices) >= 100:
            break

if prices:
    avg_price = sum(prices) / len(prices)
    print(f"  Average price (first 100): ${avg_price:.2f}")
    
    if bid_sizes and ask_sizes:
        avg_bid = sum(bid_sizes) / len(bid_sizes)
        avg_ask = sum(ask_sizes) / len(ask_sizes)
        
        print(f"\nTypical liquidity in local currency:")
        print(f"  Bid side: ${avg_bid * avg_price:,.0f} (qty: {avg_bid:,.0f} @ ${avg_price:.2f})")
        print(f"  Ask side: ${avg_ask * avg_price:,.0f} (qty: {avg_ask:,.0f} @ ${avg_price:.2f})")
        print(f"\nCurrent threshold: $13,000")
        print(f"Quote size: 65,000 shares")
        
        # Calculate how often liquidity meets threshold
        meets_threshold_bid = sum(1 for size in bid_sizes if size * avg_price >= 13000)
        meets_threshold_ask = sum(1 for size in ask_sizes if size * avg_price >= 13000)
        
        print(f"\nLiquidity threshold met:")
        print(f"  Bid side: {meets_threshold_bid}/{len(bid_sizes)} = {meets_threshold_bid/len(bid_sizes)*100:.1f}%")
        print(f"  Ask side: {meets_threshold_ask}/{len(ask_sizes)} = {meets_threshold_ask/len(ask_sizes)*100:.1f}%")

wb.close()
