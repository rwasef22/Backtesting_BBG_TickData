"""
Analyze historical data to determine optimal quote sizes for each security.
Calculates 10% of average bid/ask size between 12:00-14:00 for first 5 trading days.
"""
import openpyxl
import json
from datetime import datetime, time
from collections import defaultdict

# Load existing config
with open('configs/mm_config.json', 'r') as f:
    config = json.load(f)

# Open Excel file
wb = openpyxl.load_workbook('data/raw/TickData.xlsx', read_only=True, data_only=True)

results = {}

for sheet_name in wb.sheetnames:
    print(f"\nAnalyzing {sheet_name}...")
    ws = wb[sheet_name]
    
    # Extract security name (remove " UH Equity" or " DH Equity" suffix)
    security = sheet_name.replace(' UH Equity', '').replace(' DH Equity', '')
    
    # Collect data: date -> {bid_sizes: [], ask_sizes: []}
    daily_data = defaultdict(lambda: {'bid_sizes': [], 'ask_sizes': []})
    
    # Process rows - stop after collecting 5 days of data
    days_found = set()
    row_count = 0
    max_rows = 200000  # Limit to prevent memory issues
    
    for row in ws.iter_rows(min_row=4, max_row=max_rows, values_only=True):
        row_count += 1
        if not row[0] or not row[1]:  # Need timestamp and type
            continue
            
        # Parse timestamp
        dt = row[0]
        if not isinstance(dt, datetime):
            continue
        
        date_str = dt.date().isoformat()
        days_found.add(date_str)
        
        # Stop after collecting 5 days worth of data
        if len(days_found) > 5:
            break
        
        # Filter: 12:00 to 14:00
        if not (time(12, 0) <= dt.time() <= time(14, 0)):
            continue
        
        row_type = row[1]  # 'ASK', 'BID', or 'TRADE'
        size = row[3] if len(row) > 3 and row[3] else 0
        
        if size <= 0:
            continue
        
        if row_type == 'BID':
            daily_data[date_str]['bid_sizes'].append(size)
        elif row_type == 'ASK':
            daily_data[date_str]['ask_sizes'].append(size)
    
    print(f"  Processed {row_count:,} rows, found {len(days_found)} days")
    
    # Get first 5 trading days
    sorted_dates = sorted(daily_data.keys())[:5]
    
    if len(sorted_dates) < 5:
        print(f"  Warning: Only {len(sorted_dates)} trading days available")
    
    if len(sorted_dates) == 0:
        print(f"  No trading data found for {security}")
        continue
    
    # Calculate average bid/ask size across first 5 days
    all_bid_sizes = []
    all_ask_sizes = []
    
    for date in sorted_dates:
        all_bid_sizes.extend(daily_data[date]['bid_sizes'])
        all_ask_sizes.extend(daily_data[date]['ask_sizes'])
    
    if not all_bid_sizes and not all_ask_sizes:
        print(f"  No valid bid/ask data found for {security}")
        continue
    
    # Average bid/ask size
    avg_bid = sum(all_bid_sizes) / len(all_bid_sizes) if all_bid_sizes else 0
    avg_ask = sum(all_ask_sizes) / len(all_ask_sizes) if all_ask_sizes else 0
    avg_size = (avg_bid + avg_ask) / 2 if (avg_bid + avg_ask) > 0 else 0
    
    # Quote size = 10% of average
    quote_size = int(avg_size * 0.1)
    
    # Round to nearest 100 for cleaner values
    quote_size = max(100, round(quote_size / 100) * 100)
    
    print(f"  {security}:")
    print(f"    Trading days analyzed: {len(sorted_dates)}")
    print(f"    Avg bid size: {avg_bid:,.0f}")
    print(f"    Avg ask size: {avg_ask:,.0f}")
    print(f"    Combined avg: {avg_size:,.0f}")
    print(f"    Recommended quote_size (10%): {quote_size:,}")
    
    results[security] = {
        'quote_size': quote_size,
        'avg_bid_size': round(avg_bid, 2),
        'avg_ask_size': round(avg_ask, 2),
        'days_analyzed': len(sorted_dates)
    }

wb.close()

# Update config with new quote sizes, preserving other parameters
print("\n" + "="*70)
print("UPDATED CONFIGURATION")
print("="*70)

for security, data in results.items():
    if security in config:
        # Update existing entry
        config[security]['quote_size'] = data['quote_size']
    else:
        # Create new entry with defaults
        config[security] = {
            'quote_size': data['quote_size'],
            'refill_interval_sec': 180,
            'max_position': 100000,
            'max_notional': 1500000,
            'min_local_currency_before_quote': 25000
        }
    
    print(f"\n{security}:")
    print(f"  quote_size: {config[security]['quote_size']:,}")
    print(f"  (based on avg bid/ask: {data['avg_bid_size']:,.0f} / {data['avg_ask_size']:,.0f})")

# Save updated config
with open('configs/mm_config.json', 'w') as f:
    json.dump(config, f, indent=2)

print("\n" + "="*70)
print("Configuration saved to configs/mm_config.json")
print("="*70)
