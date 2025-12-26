import openpyxl
import json
from datetime import datetime, time
from collections import defaultdict

# Load config
with open('configs/mm_config.json', 'r') as f:
    config_data = json.load(f)

security = 'ADNOCGAS'
config = config_data[security]

output_file = 'output/adnocgas_gap_analysis.txt'
with open(output_file, 'w') as out:
    out.write(f"{'='*80}\n")
    out.write(f"ADNOCGAS Diagnostic Backtest\n")
    out.write(f"{'='*80}\n\n")
    out.write(f"Configuration:\n")
    out.write(f"  Quote Size: {config['quote_size']:,}\n")
    out.write(f"  Max Position: {config['max_position']:,}\n")
    out.write(f"  Min Local Currency: {config['min_local_currency_before_quote']:,}\n")
    out.write(f"  Refill Interval: {config['refill_interval_sec']}s\n")
    out.write(f"  Max Notional: {config['max_notional']:,}\n\n")

print(f"Starting ADNOCGAS diagnostic backtest...")
print(f"Output will be saved to: {output_file}")

# Track daily statistics
daily_stats = defaultdict(lambda: {
    'rows': 0,
    'bids': 0,
    'asks': 0,
    'trades': 0,
    'liquidity_checks': 0,
    'liquidity_pass': 0,
    'liquidity_fail': 0,
    'in_opening_auction': 0,
    'in_closing_auction': 0,
    'in_silent_period': 0,
    'quote_attempts': 0,
    'quotes_sent': 0,
    'first_valid_time': None,
    'last_valid_time': None,
    'avg_bid_value': [],
    'avg_ask_value': []
})

# Time windows
OPENING_START = time(9, 30)
OPENING_END = time(10, 0)
SILENT_START = time(10, 0)
SILENT_END = time(10, 5)
CLOSING_START = time(14, 45)
CLOSING_END = time(15, 0)

def is_in_opening_auction(ts):
    t = ts.time()
    return OPENING_START <= t < OPENING_END

def is_in_closing_auction(ts):
    t = ts.time()
    return CLOSING_START <= t < CLOSING_END

def is_in_silent_period(ts):
    t = ts.time()
    return SILENT_START <= t < SILENT_END

print("Loading Excel file...")
wb = openpyxl.load_workbook('data/raw/TickData.xlsx', read_only=False, data_only=True)

# Find ADNOCGAS sheet
sheet_name = None
for name in wb.sheetnames:
    if 'ADNOCGAS' in name.upper():
        sheet_name = name
        break

if not sheet_name:
    print("ERROR: ADNOCGAS sheet not found!")
    wb.close()
    exit(1)

print(f"Found sheet: {sheet_name}")
ws = wb[sheet_name]

# Read header (row 3)
header_row = 3
header = [cell.value for cell in ws[header_row]]
dates_col = header.index('Dates')
type_col = header.index('Type')
price_col = header.index('Price')
size_col = header.index('Size')

print("Processing data rows...")
current_date = None
bid_price = None
bid_size = None
ask_price = None
ask_size = None
last_quote_time = None
row_count = 0

for row in ws.iter_rows(min_row=header_row+1, values_only=True):
    row_count += 1
    if row_count % 10000 == 0:
        print(f"  Processed {row_count:,} rows...")
    
    date_val = row[dates_col]
    if not isinstance(date_val, datetime):
        continue
    
    date = date_val.date()
    
    # New day
    if date != current_date:
        current_date = date
        bid_price = None
        bid_size = None
        ask_price = None
        ask_size = None
        last_quote_time = None
    
    daily_stats[date]['rows'] += 1
    
    row_type = row[type_col]
    price = row[price_col]
    size = row[size_col]
    
    # Skip invalid prices
    if price is None or price <= 0:
        continue
    
    # Update orderbook
    if row_type == 'BID':
        bid_price = price
        bid_size = size
        daily_stats[date]['bids'] += 1
    elif row_type == 'ASK':
        ask_price = price
        ask_size = size
        daily_stats[date]['asks'] += 1
    elif row_type == 'TRADE':
        daily_stats[date]['trades'] += 1
    
    # Check if we're in a valid trading window
    if is_in_opening_auction(date_val):
        daily_stats[date]['in_opening_auction'] += 1
        continue
    
    if is_in_closing_auction(date_val):
        daily_stats[date]['in_closing_auction'] += 1
        continue
    
    if is_in_silent_period(date_val):
        daily_stats[date]['in_silent_period'] += 1
        continue
    
    # Track first/last valid trading time
    if daily_stats[date]['first_valid_time'] is None:
        daily_stats[date]['first_valid_time'] = date_val.time()
    daily_stats[date]['last_valid_time'] = date_val.time()
    
    # Check if we can quote (simulate strategy logic)
    if bid_price and ask_price and bid_size and ask_size:
        daily_stats[date]['liquidity_checks'] += 1
        
        # Check liquidity threshold
        bid_value = bid_price * bid_size
        ask_value = ask_price * ask_size
        
        daily_stats[date]['avg_bid_value'].append(bid_value)
        daily_stats[date]['avg_ask_value'].append(ask_value)
        
        if (bid_value >= config['min_local_currency_before_quote'] and 
            ask_value >= config['min_local_currency_before_quote']):
            daily_stats[date]['liquidity_pass'] += 1
            
            # Check refill interval
            if last_quote_time is None:
                # First quote of the day
                daily_stats[date]['quote_attempts'] += 1
                daily_stats[date]['quotes_sent'] += 1
                last_quote_time = date_val
            else:
                time_since_last = (date_val - last_quote_time).total_seconds()
                if time_since_last >= config['refill_interval_sec']:
                    daily_stats[date]['quote_attempts'] += 1
                    daily_stats[date]['quotes_sent'] += 1
                    last_quote_time = date_val
                else:
                    daily_stats[date]['quote_attempts'] += 1
        else:
            daily_stats[date]['liquidity_fail'] += 1

wb.close()
print(f"Finished processing {row_count:,} rows")

# Write analysis to file
with open(output_file, 'a') as out:
    out.write(f"{'='*80}\n")
    out.write("Daily Analysis - Non-Trading Days\n")
    out.write(f"{'='*80}\n\n")
    
    trading_days = []
    no_trading_days = []
    
    for date in sorted(daily_stats.keys()):
        stats = daily_stats[date]
        
        if stats['quotes_sent'] > 0:
            trading_days.append(date)
        else:
            no_trading_days.append(date)
            
            # Detailed analysis for non-trading days
            out.write(f"{date} - NO TRADING\n")
            out.write(f"  Total rows: {stats['rows']:,}\n")
            out.write(f"  Bids: {stats['bids']:,}, Asks: {stats['asks']:,}, Trades: {stats['trades']:,}\n")
            out.write(f"  Opening auction events: {stats['in_opening_auction']:,}\n")
            out.write(f"  Silent period events: {stats['in_silent_period']:,}\n")
            out.write(f"  Closing auction events: {stats['in_closing_auction']:,}\n")
            
            if stats['first_valid_time']:
                out.write(f"  Valid trading window: {stats['first_valid_time']} - {stats['last_valid_time']}\n")
            
            out.write(f"  Liquidity checks: {stats['liquidity_checks']:,}\n")
            out.write(f"    - Passed: {stats['liquidity_pass']:,}\n")
            out.write(f"    - Failed: {stats['liquidity_fail']:,}\n")
            
            if stats['avg_bid_value']:
                avg_bid = sum(stats['avg_bid_value']) / len(stats['avg_bid_value'])
                avg_ask = sum(stats['avg_ask_value']) / len(stats['avg_ask_value'])
                out.write(f"  Avg bid value: ${avg_bid:,.2f}\n")
                out.write(f"  Avg ask value: ${avg_ask:,.2f}\n")
                out.write(f"  Min threshold: ${config['min_local_currency_before_quote']:,}\n")
            
            out.write(f"  Quote attempts: {stats['quote_attempts']:,}\n")
            
            # Identify the main reason for no trading
            reasons = []
            if stats['rows'] == 0:
                reasons.append("No data for this day")
            elif stats['bids'] == 0 or stats['asks'] == 0:
                reasons.append("Missing bid/ask data")
            elif stats['liquidity_checks'] == 0:
                reasons.append("No valid bid/ask pairs during trading window")
            elif stats['liquidity_fail'] > 0 and stats['liquidity_pass'] == 0:
                reasons.append("Insufficient liquidity (below threshold)")
            elif stats['quote_attempts'] > 0 and stats['quotes_sent'] == 0:
                reasons.append("Refill interval not met")
            else:
                reasons.append("Unknown - need deeper investigation")
            
            out.write(f"  PRIMARY REASON: {', '.join(reasons)}\n\n")
    
    out.write(f"\n{'='*80}\n")
    out.write("Summary\n")
    out.write(f"{'='*80}\n")
    out.write(f"Total unique dates in data: {len(daily_stats)}\n")
    out.write(f"Days with trading: {len(trading_days)}\n")
    out.write(f"Days without trading: {len(no_trading_days)}\n")
    out.write(f"Trading coverage: {len(trading_days)/len(daily_stats)*100:.1f}%\n\n")
    
    out.write(f"Sample days with trading:\n")
    for date in trading_days[:10]:
        stats = daily_stats[date]
        out.write(f"  {date}: {stats['quotes_sent']:,} quotes, "
                  f"{stats['liquidity_pass']:,}/{stats['liquidity_checks']:,} liquidity checks passed\n")
    if len(trading_days) > 10:
        out.write(f"  ... and {len(trading_days)-10} more days\n")
    
    out.write(f"\n{'='*80}\n")
    out.write(f"Analysis complete. Found {len(no_trading_days)} non-trading days.\n")
    out.write(f"{'='*80}\n")

print(f"\n{'='*60}")
print(f"Analysis complete!")
print(f"Results saved to: {output_file}")
print(f"{'='*60}")
print(f"Total dates: {len(daily_stats)}")
print(f"Trading days: {len(trading_days)}")
print(f"Non-trading days: {len(no_trading_days)}")
print(f"Coverage: {len(trading_days)/len(daily_stats)*100:.1f}%")
