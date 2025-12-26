import openpyxl

wb = openpyxl.load_workbook('data/raw/TickData.xlsx', read_only=True, data_only=True)
ws = wb[wb.sheetnames[0]]

# Read first 5 rows
print(f"Sheet: {wb.sheetnames[0]}")
for i, row in enumerate(ws.iter_rows(min_row=1, max_row=5, values_only=True), 1):
    print(f"Row {i}: {row}")

wb.close()
