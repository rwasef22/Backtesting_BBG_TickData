import openpyxl
import json
from datetime import datetime, time
from collections import defaultdict

# Load config
with open('configs/mm_config.json', 'r') as f:
    config_data = json.load(f)

security = 'ADNOCGAS'
config = config_data[security]

print(f"\n{'='*80}")
print(f"ADNOCGAS Diagnostic Backtest")
print(f"{'='*80}")
print(f"\nConfiguration:")
print(f"  Quote Size: {config['quote_size']:,}")
print(f"  Max Position: {config['max_position']:,}")
print(f"  Min Local Currency: {config['min_local_currency_before_quote']:,}")
print(f"  Refill Interval: {config['refill_interval_sec']}s")
print(f"  Max Notional: {config['max_notional']:,}")

# Track daily statistics
daily_stats = defaultdict(lambda: {
    'rows': 0,
    'bids': 0,
    'asks': 0,
    'trades': 0,
    'liquidity_checks': 0,
    'liquidity_pass': 0,
    'liquidity_fail': 0,
    'in_opening_auction': 0,
    'in_closing_auction': 0,
    'in_silent_period': 0,
    'quote_attempts': 0,
    'quotes_sent': 0,
    'first_valid_time': None,
    'last_valid_time': None
})

# Time windows
OPENING_START = time(9, 30)
OPENING_END = time(10, 0)
SILENT_START = time(10, 0)
SILENT_END = time(10, 5)
CLOSING_START = time(14, 45)
CLOSING_END = time(15, 0)

def is_in_opening_auction(ts):
    t = ts.time()
    return OPENING_START <= t < OPENING_END

def is_in_closing_auction(ts):
    t = ts.time()
    return CLOSING_START <= t < CLOSING_END

def is_in_silent_period(ts):
    t = ts.time()
    return SILENT_START <= t < SILENT_END

print("\nLoading data...")
wb = openpyxl.load_workbook('data/raw/TickData.xlsx', read_only=False, data_only=True)

# Find ADNOCGAS sheet
sheet_name = None
for name in wb.sheetnames:
    if 'ADNOCGAS' in name.upper():
        sheet_name = name
        break

if not sheet_name:
    print("ERROR: ADNOCGAS sheet not found!")
    wb.close()
    exit(1)

print(f"Found sheet: {sheet_name}")
ws = wb[sheet_name]

# Read header (row 3)
header_row = 3
header = [cell.value for cell in ws[header_row]]
dates_col = header.index('Dates')
type_col = header.index('Type')
price_col = header.index('Price')
size_col = header.index('Size')

print("\nProcessing data...")
current_date = None
bid_price = None
bid_size = None
ask_price = None
ask_size = None
last_quote_time = None

for i, row in enumerate(ws.iter_rows(min_row=header_row+1, values_only=True), start=header_row+1):
    if i % 50000 == 0:
        print(f"  Processed {i-header_row:,} rows...")
    
    date_val = row[dates_col]
    if not isinstance(date_val, datetime):
        continue
    
    date = date_val.date()
    
    # New day
    if date != current_date:
        current_date = date
        bid_price = None
        bid_size = None
        ask_price = None
        ask_size = None
        last_quote_time = None
    
    daily_stats[date]['rows'] += 1
    
    row_type = row[type_col]
    price = row[price_col]
    size = row[size_col]
    
    # Skip invalid prices
    if price is None or price <= 0:
        continue
    
    # Update orderbook
    if row_type == 'BID':
        bid_price = price
        bid_size = size
        daily_stats[date]['bids'] += 1
    elif row_type == 'ASK':
        ask_price = price
        ask_size = size
        daily_stats[date]['asks'] += 1
    elif row_type == 'TRADE':
        daily_stats[date]['trades'] += 1
    
    # Check if we're in a valid trading window
    if is_in_opening_auction(date_val):
        daily_stats[date]['in_opening_auction'] += 1
        continue
    
    if is_in_closing_auction(date_val):
        daily_stats[date]['in_closing_auction'] += 1
        continue
    
    if is_in_silent_period(date_val):
        daily_stats[date]['in_silent_period'] += 1
        continue
    
    # Track first/last valid trading time
    if daily_stats[date]['first_valid_time'] is None:
        daily_stats[date]['first_valid_time'] = date_val.time()
    daily_stats[date]['last_valid_time'] = date_val.time()
    
    # Check if we can quote (simulate strategy logic)
    if bid_price and ask_price and bid_size and ask_size:
        daily_stats[date]['liquidity_checks'] += 1
        
        # Check liquidity threshold
        bid_value = bid_price * bid_size
        ask_value = ask_price * ask_size
        
        if (bid_value >= config['min_local_currency_before_quote'] and 
            ask_value >= config['min_local_currency_before_quote']):
            daily_stats[date]['liquidity_pass'] += 1
            
            # Check refill interval
            if last_quote_time is None:
                # First quote of the day
                daily_stats[date]['quote_attempts'] += 1
                daily_stats[date]['quotes_sent'] += 1
                last_quote_time = date_val
            else:
                time_since_last = (date_val - last_quote_time).total_seconds()
                if time_since_last >= config['refill_interval_sec']:
                    daily_stats[date]['quote_attempts'] += 1
                    daily_stats[date]['quotes_sent'] += 1
                    last_quote_time = date_val
                else:
                    daily_stats[date]['quote_attempts'] += 1
        else:
            daily_stats[date]['liquidity_fail'] += 1

wb.close()

# Analyze results
print(f"\n{'='*80}")
print("Daily Analysis")
print(f"{'='*80}")

trading_days = []
no_trading_days = []

for date in sorted(daily_stats.keys()):
    stats = daily_stats[date]
    
    if stats['quotes_sent'] > 0:
        trading_days.append(date)
    else:
        no_trading_days.append(date)
        
        # Detailed analysis for non-trading days
        print(f"\n{date} - NO TRADING")
        print(f"  Total rows: {stats['rows']:,}")
        print(f"  Bids: {stats['bids']:,}, Asks: {stats['asks']:,}, Trades: {stats['trades']:,}")
        print(f"  Opening auction events: {stats['in_opening_auction']:,}")
        print(f"  Silent period events: {stats['in_silent_period']:,}")
        print(f"  Closing auction events: {stats['in_closing_auction']:,}")
        
        if stats['first_valid_time']:
            print(f"  Valid trading window: {stats['first_valid_time']} - {stats['last_valid_time']}")
        
        print(f"  Liquidity checks: {stats['liquidity_checks']:,}")
        print(f"    - Passed: {stats['liquidity_pass']:,}")
        print(f"    - Failed: {stats['liquidity_fail']:,}")
        print(f"  Quote attempts: {stats['quote_attempts']:,}")
        
        # Identify the main reason for no trading
        reasons = []
        if stats['bids'] == 0 or stats['asks'] == 0:
            reasons.append("Missing bid/ask data")
        if stats['liquidity_fail'] > 0 and stats['liquidity_pass'] == 0:
            reasons.append("Insufficient liquidity")
        if stats['quote_attempts'] > 0 and stats['quotes_sent'] == 0:
            reasons.append("Refill interval not met")
        if stats['liquidity_checks'] == 0:
            reasons.append("No valid bid/ask pairs")
        if not reasons:
            reasons.append("Unknown - need deeper investigation")
        
        print(f"  Primary reason: {', '.join(reasons)}")

print(f"\n{'='*80}")
print("Summary")
print(f"{'='*80}")
print(f"Total unique dates in data: {len(daily_stats)}")
print(f"Days with trading: {len(trading_days)}")
print(f"Days without trading: {len(no_trading_days)}")
print(f"Trading coverage: {len(trading_days)/len(daily_stats)*100:.1f}%")

print(f"\nDays with trading:")
for date in trading_days[:10]:
    stats = daily_stats[date]
    print(f"  {date}: {stats['quotes_sent']:,} quotes, "
          f"{stats['liquidity_pass']:,}/{stats['liquidity_checks']:,} liquidity checks passed")
if len(trading_days) > 10:
    print(f"  ... and {len(trading_days)-10} more days")

print(f"\n{'='*80}")
print(f"Analysis complete. Found {len(no_trading_days)} non-trading days.")
print(f"{'='*80}")
