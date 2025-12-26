import sys
import openpyxl
import json
from datetime import datetime, time
import pandas as pd

sys.path.insert(0, 'src')
from orderbook import OrderBook
from market_making_strategy import MarketMakingStrategy
from mm_handler import MarketMakingHandler

# Load config
with open('configs/mm_config.json', 'r') as f:
    config_data = json.load(f)

security = 'ADNOCGAS'
config = config_data[security]

print(f"\n{'='*80}")
print(f"ADNOCGAS Backtest with Detailed Logging")
print(f"{'='*80}")
print(f"Config: quote_size={config['quote_size']}, max_position={config['max_position']}")
print(f"        min_currency={config['min_local_currency_before_quote']}, refill={config['refill_interval_sec']}s")

# Load Excel
print("\nLoading Excel file...")
wb = openpyxl.load_workbook('data/raw/TickData.xlsx', read_only=True, data_only=True)

# Find sheet
sheet_name = None
for name in wb.sheetnames:
    if 'ADNOCGAS' in name.upper():
        sheet_name = name
        break

if not sheet_name:
    print("ERROR: ADNOCGAS sheet not found!")
    wb.close()
    exit(1)

print(f"Processing sheet: {sheet_name}")
security_name = sheet_name.replace(' UH Equity', '')
print(f"Security name: {security_name}")

# Initialize strategy components
orderbook = OrderBook()
strategy = MarketMakingStrategy(
    quote_size=config['quote_size'],
    max_position=config['max_position'],
    refill_interval_sec=config['refill_interval_sec'],
    max_notional=config['max_notional']
)
handler = MarketMakingHandler(
    security=security_name,
    orderbook=orderbook,
    strategy=strategy,
    min_local_currency_before_quote=config['min_local_currency_before_quote']
)

# Process data
ws = wb[sheet_name]
header_row = 3
header = [cell.value for cell in ws[header_row]]

chunk_size = 100000
chunk = []
current_date = None
daily_summary = {}

print("\nProcessing data...")
for i, row in enumerate(ws.iter_rows(min_row=header_row+1, values_only=True), start=1):
    if i % 50000 == 0:
        print(f"  Processed {i:,} rows, {len(daily_summary)} days analyzed")
    
    chunk.append(row)
    
    if len(chunk) >= chunk_size:
        df = pd.DataFrame(chunk, columns=header)
        
        # Check for date changes within chunk
        df['Dates'] = pd.to_datetime(df['Dates'])
        dates_in_chunk = df['Dates'].dt.date.unique()
        
        for date in dates_in_chunk:
            if date != current_date:
                # Save previous day summary
                if current_date is not None:
                    daily_summary[current_date] = {
                        'position': strategy.position,
                        'pnl': strategy.cumulative_pnl,
                        'trades': handler.trade_count,
                        'orderbook_valid': orderbook.bid_price is not None and orderbook.ask_price is not None
                    }
                
                current_date = date
        
        handler.process_chunk(df)
        chunk = []

# Process remaining chunk
if chunk:
    df = pd.DataFrame(chunk, columns=header)
    handler.process_chunk(df)
    if current_date is not None:
        daily_summary[current_date] = {
            'position': strategy.position,
            'pnl': strategy.cumulative_pnl,
            'trades': handler.trade_count,
            'orderbook_valid': orderbook.bid_price is not None and orderbook.ask_price is not None
        }

wb.close()

print(f"\n{'='*80}")
print("Daily Summary Analysis")
print(f"{'='*80}")

# Analyze trading days
trading_days = 0
no_trading_days = 0
prev_trades = 0

for date in sorted(daily_summary.keys()):
    info = daily_summary[date]
    day_trades = info['trades'] - prev_trades
    
    if day_trades > 0:
        trading_days += 1
    else:
        no_trading_days += 1
        print(f"{date}: NO TRADES - pos={info['position']}, pnl=${info['pnl']:.2f}, "
              f"orderbook_valid={info['orderbook_valid']}")
    
    prev_trades = info['trades']

print(f"\n{'='*80}")
print(f"Summary:")
print(f"  Total days: {len(daily_summary)}")
print(f"  Days with trades: {trading_days}")
print(f"  Days without trades: {no_trading_days}")
print(f"  Coverage: {trading_days/len(daily_summary)*100:.1f}%")
print(f"  Final position: {strategy.position}")
print(f"  Final P&L: ${strategy.cumulative_pnl:.2f}")
print(f"  Total trades: {handler.trade_count}")
print(f"{'='*80}")

# Check handler state
print(f"\nHandler State:")
print(f"  Last refill time: {handler.last_refill_time}")
print(f"  Trade count: {handler.trade_count}")

print(f"\nStrategy State:")
print(f"  Position: {strategy.position}")
print(f"  Cumulative P&L: ${strategy.cumulative_pnl:.2f}")
print(f"  Last refill: {strategy.last_refill_time}")

print(f"\nOrderBook State:")
print(f"  Bid: {orderbook.bid_price} x {orderbook.bid_size}")
print(f"  Ask: {orderbook.ask_price} x {orderbook.ask_size}")
