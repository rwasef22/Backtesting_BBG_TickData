"""Check actual values in ADNOCGAS sheet - first 100 rows."""
import openpyxl

print("Loading TickData.xlsx...")
wb = openpyxl.load_workbook('data/raw/TickData.xlsx', read_only=True, data_only=True)
sheet = wb['ADNOCGAS UH Equity']

print("\nFirst row (header check):")
for row in sheet.iter_rows(min_row=1, max_row=1, values_only=True):
    print(f"  {row}")

print("\nFirst 20 data rows:")
print(f"{'Row':<5} {'Col0':<20} {'Col1':<12} {'Col2':<12} {'Col3':<12}")
print("="*70)

for i, row in enumerate(sheet.iter_rows(min_row=2, max_row=22, values_only=True)):
    col0 = str(row[0])[:19] if row[0] is not None else "None"
    col1 = str(row[1]) if row[1] is not None else "None"  
    col2 = str(row[2]) if row[2] is not None else "None"
    col3 = str(row[3]) if row[3] is not None else "None"
    
    print(f"{i+2:<5} {col0:<20} {col1:<12} {col2:<12} {col3:<12}")

wb.close()

print("\n" + "="*60)
print("Check if event_type is 'bid'/'ask'/'trade' (lowercase) or 'BID'/'ASK'/'TRADE' (uppercase)")
